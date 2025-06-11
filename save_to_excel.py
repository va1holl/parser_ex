from openpyxl import Workbook, load_workbook


def create_template(headers: list[str], filename: str = "template.xlsx"):
    wb = Workbook()
    ws = wb.active

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = header

    wb.save(filename)

def save_to_excel(products_dict: dict, template_path='template.xlsx', output_path='result.xlsx'):
    create_template(list(products_dict.keys()))

    wb = load_workbook(template_path)
    ws = wb.active

    row = ws.max_row + 1

    for col, key in enumerate(products_dict.keys(), start=1):
        value = products_dict[key]
        if isinstance(value, list):
            ws.cell(row=row, column=col).value = ', '.join(value)
        else:
            ws.cell(row=row, column=col).value = value

    wb.save(output_path)